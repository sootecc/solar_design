"""
결과 출력 유틸리티 (Excel, PDF 등)
"""
import io
from typing import List
from core.models import DesignResult, ProjectInfo


def results_to_dataframe(results: List[DesignResult]):
    """DesignResult 목록을 pandas DataFrame으로 변환"""
    try:
        import pandas as pd
    except ImportError:
        return None

    rows = []
    for r in results:
        for m in r.자재목록:
            rows.append({
                "배열번호": r.배열번호,
                "자재명": m["자재명"],
                "규격": m["규격"],
                "단위": m["단위"],
                "수량": m["수량"],
                "단위무게(kg)": m["단위무게_kg"],
                "총무게(kg)": m["총무게_kg"],
            })

    return pd.DataFrame(rows)


def export_to_excel(
    results: List[DesignResult],
    project_info: ProjectInfo,
    generation_info: dict = None,
) -> bytes:
    """
    Excel 파일로 내보내기
    Returns: bytes (Excel 파일 내용)
    """
    try:
        import pandas as pd
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    wb = openpyxl.Workbook()

    # ---- 발전설비 중량계산표 시트 ----
    ws = wb.active
    ws.title = "발전설비중량계산표"

    # 스타일 정의
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    sub_header_fill = PatternFill("solid", fgColor="BDD7EE")
    sub_header_font = Font(bold=True, size=9)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 제목
    ws.merge_cells("B2:I2")
    ws["B2"] = "발전설비 중량계산표"
    ws["B2"].font = Font(bold=True, size=14)
    ws["B2"].alignment = center

    # 프로젝트 정보
    ws["B3"] = f"현장명: {project_info.현장명}"
    ws["E3"] = f"설계자: {project_info.설계자}"
    ws["G3"] = f"설계일자: {project_info.설계일자}"

    # 헤더
    headers = ["번호", "자재명", "규격", "설치위치", "단위", "단위무게(kg)", "수량", "총무게(kg)"]
    for col, h in enumerate(headers, 2):
        cell = ws.cell(row=5, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = [6, 20, 30, 15, 6, 12, 8, 12][col-2]

    # 데이터 행
    row = 6
    num = 1
    total_weight = 0.0
    for result in results:
        # 배열 헤더
        ws.merge_cells(f"B{row}:I{row}")
        ws[f"B{row}"] = f"▶ 배열 {result.배열번호} (용량: {result.총용량_kW}kW / 모듈: {result.모듈수량}장)"
        ws[f"B{row}"].fill = sub_header_fill
        ws[f"B{row}"].font = sub_header_font
        row += 1

        for m in result.자재목록:
            ws.cell(row, 2, num).alignment = center
            ws.cell(row, 3, m["자재명"])
            ws.cell(row, 4, m["규격"])
            ws.cell(row, 5, "")
            ws.cell(row, 6, m["단위"]).alignment = center
            ws.cell(row, 7, m["단위무게_kg"]).alignment = center
            ws.cell(row, 8, m["수량"]).alignment = center
            ws.cell(row, 9, m["총무게_kg"]).alignment = center
            for col in range(2, 10):
                ws.cell(row, col).border = border
            total_weight += m["총무게_kg"]
            num += 1
            row += 1

    # 합계
    ws.cell(row, 2, "합  계").font = Font(bold=True)
    ws.merge_cells(f"B{row}:H{row}")
    ws.cell(row, 9, round(total_weight, 2)).font = Font(bold=True)
    ws.cell(row, 9).alignment = center

    # ---- 발전량 추정 시트 ----
    if generation_info:
        ws2 = wb.create_sheet("발전량추정")
        ws2["B2"] = "발전량 추정"
        ws2["B2"].font = Font(bold=True, size=14)

        gen_rows = [
            ("설비용량", f"{generation_info['설비용량_kW']} kW"),
            ("일평균 일조시간", f"{generation_info['일평균일조시간_h']} h/일"),
            ("성능계수(PR)", f"{generation_info['성능계수_PR']}"),
            ("일 발전량", f"{generation_info['일발전량_kWh']} kWh"),
            ("월 발전량", f"{generation_info['월발전량_kWh']:,.0f} kWh"),
            ("연간 발전량", f"{generation_info['연발전량_kWh']:,.0f} kWh"),
        ]
        for i, (label, value) in enumerate(gen_rows, 4):
            ws2.cell(i, 2, label).font = Font(bold=True)
            ws2.cell(i, 3, value)

    # 바이트로 변환
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
